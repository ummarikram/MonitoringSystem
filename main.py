#Dependencies

import cv2   # OpenCV
import numpy as np
import face_recognition # https://pypi.org/project/face-recognition/ Uses dlib's model for accurate results
import os  # For getting file paths
import openpyxl  # for manipulating/working with excel files
from scipy.spatial import distance as dist
import imutils
from openpyxl import Workbook
from datetime import datetime  # get times
import smtplib

# For Automatic Attendance
dataSetPath = 'Known Faces' # Path to the already known faces or trained DataSet
studentImages = []  # List for storing images
studentNames = []  # List for stroring Names
marked = [] # List of Marked Attendance wrt to studentNames ; 1 OR 0
Date = datetime.today().strftime('%d-%m-%Y') # Today's Date
listOfFileNames = os.listdir(dataSetPath) # Store all fileNames in a list

# For Activeness Detection
Counter = 0
TimeSlept = 0
TotalTime = 0
BlinkTime = 5
MinEyeAspectRatio = 0.25

# For Sending Email
studentEmails = []
SenderGmailUserName = 'youremail@gmail.com'
SenderGmailPassWord = 'password'

# Loop through the list of file names
for imageName in listOfFileNames:
    relativePath = f'{dataSetPath}/{imageName}'  # Extract the relative path of the current Image
    currentImage = cv2.imread(relativePath)  # Capture the Current Image via its relative Path
    studentImages.append(currentImage)  # Add it to the list of Student Images
    studentNames.append(str(imageName).split('.')[0])   # Add the Student Name by splitting using '.' delimeter
    marked.append(0)

# Processes the list of Images and encodes them as per their features/properties in a list
# Reference : https://pypi.org/project/face-recognition/
def getEncodings(images):
    
    encodedList = []

    for image in images:
        currentImage = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(currentImage)[0]
        encodedList.append(encode)

    return encodedList

# Mark Behaviour for the specific Student in an excel sheet using openpyxl
def markAttendance(name):
    
    wb = Workbook()  # Create new Workbook
    Behaviour = wb.active  # Get Active Sheet
    Behaviour.title = "BEHAVIOUR"
    Behaviour.cell(row=1, column = 1).value = 'STUDENT NAME'  # Assign top headers
    Behaviour.cell(row=1, column = 2).value = 'ATTENDANCE TIME'
    Behaviour.cell(row=1, column = 3).value = 'ACTIVENESS LEVEL'
    MaxRow = Behaviour.max_row + 1  
    MaxCol = Behaviour.max_column + 1

    NameFound = 0  # Bool to check if attendance was marked already or not

    for i in range(2, MaxRow):
        if (Behaviour.cell(row=i, column = 1).value == name):  # if found
            NameFound = 1

    if (NameFound == 0):  # if not found then add it
        Behaviour.cell(row=MaxRow, column = 1).value = name
        Behaviour.cell(row=MaxRow, column = 2).value = datetime.now().strftime('%H:%M:%S')

    # Increase column dimensions so that text is readable 
    for sheet in wb:
        sheet.column_dimensions['A'].width = 25  # Name Column
        sheet.column_dimensions['B'].width = 25  # Time Column
        sheet.column_dimensions['C'].width = 25  # Time Column

    FileName = 'Behaviour ' + Date + '.xlsx'
    wb.save(FileName)

def attendanceSystem(img):

    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)   # Convert from BGR to RGB

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    for encodedFace, faceLocation in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodedList, encodedFace)
        faceDis = face_recognition.face_distance(encodedList, encodedFace)

        matchedIndex = np.argmin(faceDis) # used to get the indices of the minimum element from an array (single-dimensional array)
                                          # or any row or column (multidimensional array) of any given array.

        # if known face found
        if matches[matchedIndex]:

            name = studentNames[matchedIndex].upper()  # Get Name
            y1, x2, y2, x1 = faceLocation # Face Location Coordinates
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4 # Location where green box will be displayed
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)  # Green Square
            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)  # White Text with Student Name

            # if attendance is not marked before
            if (marked[matchedIndex] == 0):
                markAttendance(name)
                marked[matchedIndex] = 1 # Mark Attendance
                print('Attendace Marked for ' + name)
            
            markActivenessLevel(name)
        
        

def markActivenessLevel(name):
    
    FileName = 'Behaviour ' + Date + '.xlsx'
    wb = openpyxl.load_workbook(FileName)
    
    Behaviour = wb.active  # Get Active Sheet
    
    MaxRow = Behaviour.max_row + 1  
    MaxCol = Behaviour.max_column + 1

    for i in range(2, MaxRow):
        if (Behaviour.cell(row=i, column = 1).value == name):  # if found
            if (TotalTime>0):  # Handling Division by 0 exception
                Behaviour.cell(row=i, column = MaxCol-1).value = (1.0 - (TimeSlept/TotalTime)) * 100
            else:
                Behaviour.cell(row=i, column = MaxCol-1).value = 100  

    FileName = 'Behaviour ' + Date + '.xlsx'
    wb.save(FileName)

def eyeAspectRatio(eye):

    D1 = dist.euclidean(eye[1],eye[5]) # The Euclidean Distance will tell us whether eye is close or not
    D2 = dist.euclidean(eye[2], eye[4])
    H = dist.euclidean(eye[0], eye[3])
    ear = (D1+D2)/(2.0*H)
    return ear

def activenessDetectionSystem(img):

    face_landmarks_list = face_recognition.face_landmarks(img)
    global Counter
    global TimeSlept
    global TotalTime
    TotalTime+=1

    for face_landmark in face_landmarks_list:

        leftEye = face_landmark["left_eye"]   # Get Left Eye landmarks
        rightEye = face_landmark["right_eye"]  # Get Right Eye landmarks
        leftEyePoints = np.array(leftEye)   # Convert it to numpy array
        rightEyePoints = np.array(rightEye)   
        cv2.polylines(img,[leftEyePoints],True,(255,255,0),1)  # Dray Polylines around eyes
        cv2.polylines(img, [rightEyePoints], True, (255, 255, 0), 1)
        leftEar = eyeAspectRatio(leftEye)  # Get Aspect Ratio
        rightEar = eyeAspectRatio(rightEye)
        CurrentAspectRatio = (leftEar+rightEar)/2

        # if current aspect ration is below minimum
        if CurrentAspectRatio<MinEyeAspectRatio:
            Counter+=1  # Increase eye close Count
            if Counter>=BlinkTime:  # if eye was closed for 5 secs
                cv2.putText(img,"SLEEPY!",(20,20),
                            cv2.FONT_HERSHEY_SIMPLEX,0.5,(0,0,255),1)
                TimeSlept+=1
        else:
            Counter=0  # Reset Counter

encodedList = getEncodings(studentImages)
print('Encoding Completed')

cap = cv2.VideoCapture(0)  # Initialize Webcam

while True:
    
    success, img = cap.read()   # Read Current Frame

    activenessDetectionSystem(img)  

    attendanceSystem(img)

    cv2.imshow('WEBCAM',img)  # Show Webcam

    # if escape key pressed
    if cv2.waitKey(1) == 27:
        break
    
    # if window closed
    if cv2.getWindowProperty('WEBCAM', cv2.WND_PROP_VISIBLE) < 1:
        break

cap.release()  # Release Camera & Destroy all windows
cv2.destroyAllWindows()


def sendEmail(studentEmail):
    
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.ehlo()

    server.login(SenderGmailUserName, SenderGmailPassWord)

    subject = 'BEHAVIOUR REPORT'

    body = 'Your Attendance was marked'

    msg = f"Subject: {subject}\n\n\n\n{body}"

    server.sendmail(
        SenderGmailUserName, # From
        SenderGmailUserName, # To
        msg            # Message
    )

    print('EMAIL SENT!')

    server.quit()

studentEmails.append('email@gmail.com')
#for email in studentEmails:
    #sendEmail(email)